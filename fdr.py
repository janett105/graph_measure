import pandas as pd
from statsmodels.stats.multitest import multipletests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


df = pd.read_csv("corr_new.csv")

# FDR 보정 계산
p_values = df['p_value'].values  # 'p_value'를 실제 열 이름으로 대체하세요.
reject, pvals_corrected, _, _ = multipletests(p_values, alpha=0.05, method='fdr_bh')

# 보정된 p-value를 DataFrame에 추가
df['FDR_corrected_p_value'] = pvals_corrected

# Excel 파일로 저장
output_filename = "corr_new_FDR.xlsx"
df.to_excel(output_filename, index=False)

# Excel 파일을 불러와 셀 하이라이트
wb = load_workbook(output_filename)
ws = wb.active

# 분홍색 채우기 설정
pink_fill = PatternFill(start_color='FFC0CB',
                        end_color='FFC0CB',
                        fill_type='solid')

# 조건에 맞는 셀 하이라이트
for row in range(2, ws.max_row + 1):
    _cell = ws.cell(row, df.columns.get_loc('FDR_corrected_p_value') + 1)
    if _cell.value < 0.05:
        _cell.fill = pink_fill

# 변경 사항 저장
wb.save(output_filename)
